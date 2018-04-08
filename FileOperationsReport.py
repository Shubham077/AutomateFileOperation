import os
import time
from openpyxl import Workbook
from openpyxl import load_workbook

reportfilename = time.strftime("%Y%m%d-%H%M%S") + ".xlsx"
currentpath = "C:\\Users\\Admin\\Desktop"               #path where the report file will be stored
current_column = 0
current_row = 0


def createlogfile():                    #create excel report file
    os.chdir(currentpath)
    wb = Workbook()
    wb.save(reportfilename)


def get_currentlocation(file_number):               # tells where to paste data in excel
    global current_column
    global current_row
    if current_row != file_number:
        current_column = 0              # if a next file is now operated on, the column_number will reset
    current_column = current_column+1
    current_row = file_number
    location = chr(64+current_column) + str(current_row)                # will tell exact location ie: A1 or A3 or B1
    return (location)


def reportlog(content,filenum):
    location = get_currentlocation(filenum)
    writetologfile(content,location)


def writetologfile(content,location):               # open, edit and close the excel report file
    os.chdir(currentpath)
    wb = load_workbook(reportfilename)
    ws = wb.active
    ws[location] = content
    wb.save(reportfilename)